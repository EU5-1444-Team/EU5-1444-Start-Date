#!/usr/bin/env python3
"""Convert an Excel pop file (produced by pops_to_excel.py) back to 06_pops.txt format."""

from __future__ import annotations

import json
import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

SETTINGS_PATH  = Path(__file__).resolve().parent / "settings.json"
STD_ATTR_ORDER = ["type", "size", "culture", "religion"]
# Columns we read from the sheet — geo/differs/basegame are ignored on import
IMPORT_COLS    = {"location", "type", "size", "culture", "religion", "extra"}


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

def load_settings() -> dict:
    if SETTINGS_PATH.is_file():
        try:
            s = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
            if "base_game" in s and "mod_folder" in s:
                return s
        except json.JSONDecodeError:
            pass
    return {}


def prompt_for_paths() -> dict:
    print("Settings file missing or malformed.")
    base_game  = input("Base game path (root with /game subfolder): ").strip()
    mod_folder = input("Mod folder path: ").strip()
    settings   = {"base_game": base_game, "mod_folder": mod_folder}
    SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
    SETTINGS_PATH.write_text(json.dumps(settings, indent=2), encoding="utf-8")
    return settings


# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------

def read_excel(path: Path) -> tuple[list[str], dict[str, list[OrderedDict]]]:
    """Read the Pops sheet and return (location_order, location_pops).

    location_order preserves the row order from the sheet.
    location_pops maps location -> list of attr dicts in order.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Pops" not in wb.sheetnames:
        raise ValueError(f"No 'Pops' sheet found in {path}")
    ws = wb["Pops"]

    rows = ws.iter_rows(values_only=True)
    raw_headers = next(rows)
    headers = [str(h).lower() if h is not None else "" for h in raw_headers]

    # Map column name -> index
    col = {name: idx for idx, name in enumerate(headers)}
    missing = IMPORT_COLS - set(col)
    if missing:
        raise ValueError(f"Sheet is missing required columns: {', '.join(sorted(missing))}")

    location_order: list[str] = []
    location_pops:  dict[str, list[OrderedDict]] = {}

    for row in rows:
        location = str(row[col["location"]] or "").strip()
        if not location:
            continue

        attrs: OrderedDict[str, str] = OrderedDict()

        # Standard attributes in canonical order
        for key in STD_ATTR_ORDER:
            raw = row[col[key]]
            if raw is None:
                continue
            value = str(raw).strip()
            if not value:
                continue
            if key == "size":
                try:
                    value = f"{float(value):.3f}"
                except ValueError:
                    pass
            attrs[key] = value

        # Extra attributes (format: "key=value; key=value")
        extra_raw = row[col["extra"]]
        if extra_raw:
            for part in str(extra_raw).split(";"):
                part = part.strip()
                if not part or "=" not in part:
                    continue
                k, v = part.split("=", 1)
                k, v = k.strip(), v.strip()
                if k and v:
                    attrs[k] = v

        if not attrs:
            continue

        if location not in location_pops:
            location_order.append(location)
            location_pops[location] = []
        location_pops[location].append(attrs)

    wb.close()
    return location_order, location_pops


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def normalize_size(value: str) -> str:
    try:
        return f"{float(value):.3f}"
    except Exception:
        return value


def format_pop_line(attrs: OrderedDict) -> str:
    ordered = OrderedDict()
    for key in STD_ATTR_ORDER:
        if key in attrs:
            ordered[key] = attrs[key]
    for key, value in attrs.items():
        if key not in ordered:
            ordered[key] = value
    parts = "\t".join(f"{k} = {v}" for k, v in ordered.items())
    return f"\tdefine_pop = {{ {parts} }}"


def write_pops(location_order: list[str], location_pops: dict[str, list[OrderedDict]], output_path: Path) -> None:
    lines = ["locations={", ""]
    for location in location_order:
        lines.append(f"{location} = {{")
        for attrs in location_pops.get(location, []):
            lines.append(format_pop_line(attrs))
        lines.append("}")
    lines.append("}")
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    settings = load_settings()
    if not settings:
        settings = prompt_for_paths()

    mod_folder_path = Path(settings["mod_folder"]).expanduser()

    input_path  = Path(__file__).resolve().parent / "06_pops.xlsx"
    output_path = mod_folder_path / "main_menu/setup/start/06_pops.txt"

    if len(sys.argv) >= 2: input_path  = Path(sys.argv[1])
    if len(sys.argv) >= 3: output_path = Path(sys.argv[2])

    if not input_path.exists():
        print(f"Error: input file not found: {input_path}")
        sys.exit(1)

    print(f"Reading Excel from {input_path} ...")
    location_order, location_pops = read_excel(input_path)

    total_pops = sum(len(v) for v in location_pops.values())
    print(f"Read {total_pops} pop rows across {len(location_order)} locations.")

    # Backup existing file
    if output_path.exists():
        backup = output_path.with_suffix(output_path.suffix + ".bak")
        backup.write_text(output_path.read_text(encoding="utf-8-sig"), encoding="utf-8")
        print(f"Backed up existing file to {backup.name}")

    print(f"Writing {output_path} ...")
    write_pops(location_order, location_pops, output_path)
    print("Done.")


if __name__ == "__main__":
    main()
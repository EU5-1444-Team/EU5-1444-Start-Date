#!/usr/bin/env python3
"""Convert EU5 1444 pop data to Excel with geography columns."""

from __future__ import annotations

import json
import re
import sys
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SETTINGS_PATH = Path(__file__).resolve().parent / "settings.json"
BASE_POPS_REL = Path("game/main_menu/setup/start/06_pops.txt")
STD_ATTR_ORDER = ["type", "size", "culture", "religion"]
TABLE_COLUMNS = [
    "location", "type", "size", "culture", "religion",
    "continent", "superregion", "region", "area", "province",
    "extra", "differs", "basegame",
]
GEO_COLS = {"continent", "superregion", "region", "area", "province"}
COL_WIDTHS = {
    "location": 20, "type": 14, "size": 10, "culture": 18, "religion": 18,
    "extra": 35, "continent": 16, "superregion": 20, "region": 20,
    "area": 20, "province": 22, "differs": 9, "basegame": 45,
}

_GEO_COL_IDXS    = {TABLE_COLUMNS.index(c) for c in GEO_COLS}
_SIZE_COL_IDX    = TABLE_COLUMNS.index("size")
_DIFFERS_COL_IDX = TABLE_COLUMNS.index("differs")

# Row 1 = live totals, Row 2 = headers with filter dropdowns, Row 3+ = data
_DATA_START_ROW = 3


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class GeoInfo:
    continent: str = ""
    superregion: str = ""
    region: str = ""
    area: str = ""
    province: str = ""


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

def load_settings() -> dict:
    if SETTINGS_PATH.is_file():
        try:
            s = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
            if "base_game" in s and "mod_folder" in s:
                return s
        except json.JSONDecodeError:
            pass
    return {}


def prompt_for_paths() -> dict:
    print("Settings file missing or malformed.")
    base_game = input("Base game path (root with /game subfolder): ").strip()
    mod_folder = input("Mod folder path: ").strip()
    settings = {"base_game": base_game, "mod_folder": mod_folder}
    SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
    SETTINGS_PATH.write_text(json.dumps(settings, indent=2), encoding="utf-8")
    return settings


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def normalize_size(value: str) -> str:
    try:
        return f"{float(value):.3f}"
    except Exception:
        return value


def tokenize_definitions(text: str) -> list[str]:
    cleaned = re.sub(r"#.*$", "", text, flags=re.MULTILINE)
    return re.findall(r"[{}=]|[^\s{}=]+", cleaned)


def classify_geo(path: tuple) -> GeoInfo:
    values = list(path) + [""] * 5
    return GeoInfo(*values[:5])


def _parse_geo_node(tokens, idx, name, path, location_geo):
    while idx < len(tokens):
        token = tokens[idx]
        if token == "}":
            return idx + 1
        key = token
        idx += 1
        if idx < len(tokens) and tokens[idx] == "=":
            idx += 1
            if tokens[idx] != "{":
                raise ValueError(f"Expected '{{' after {key}")
            idx = _parse_geo_node(tokens, idx + 1, key, path + (key,), location_geo)
        else:
            location_geo[key] = classify_geo(path)
    raise ValueError(f"Unclosed geography block for {name}")


def parse_definitions(path: Path) -> dict[str, GeoInfo]:
    tokens = tokenize_definitions(path.read_text(encoding="utf-8-sig"))
    idx = 0
    location_geo: dict[str, GeoInfo] = {}
    while idx < len(tokens):
        name = tokens[idx]
        idx += 1
        if idx >= len(tokens) or tokens[idx] != "=":
            raise ValueError(f"Expected '=' after {name}")
        idx += 1
        if tokens[idx] != "{":
            raise ValueError(f"Expected '{{' after {name} =")
        idx = _parse_geo_node(tokens, idx + 1, name, (name,), location_geo)
    return location_geo


def _find_matching_brace(text: str, start: int) -> int:
    depth = 0
    for i in range(start, len(text)):
        c = text[i]
        if c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return i
    raise ValueError("Unmatched brace")


def parse_pops(path: Path) -> tuple[list[str], dict[str, list[OrderedDict]]]:
    text = path.read_text(encoding="utf-8-sig")
    root_match = re.search(r"locations\s*=\s*\{", text)
    if not root_match:
        raise ValueError("Could not find locations={...} in pop file")
    root_brace = text.index("{", root_match.start())
    body = text[root_brace + 1: _find_matching_brace(text, root_brace)]

    location_order: list[str] = []
    location_pops: dict[str, list[OrderedDict]] = {}
    idx = 0
    while idx < len(body):
        m = re.search(r"([A-Za-z0-9_]+)\s*=", body[idx:])
        if not m:
            break
        location = m.group(1)
        abs_start = idx + m.start()
        brace_idx = body.index("{", abs_start + len(m.group(0)))
        block_end = _find_matching_brace(body, brace_idx)
        block_text = body[brace_idx + 1: block_end]
        pops = [
            OrderedDict(re.findall(r"([A-Za-z0-9_]+)\s*=\s*([^\s}]+)", pb))
            for pb in re.findall(r"define_pop\s*=\s*\{([^{}]*)\}", block_text)
        ]
        location_order.append(location)
        location_pops[location] = [p for p in pops if p]
        idx = block_end + 1
    return location_order, location_pops


# ---------------------------------------------------------------------------
# Row building
# ---------------------------------------------------------------------------

def build_rows(
    location_order: list[str],
    location_pops: dict[str, list[OrderedDict]],
    location_geo: dict[str, GeoInfo],
    base_location_pops: dict[str, list[OrderedDict]],
) -> list[tuple]:
    rows = []
    _empty_geo = GeoInfo()
    for location in location_order:
        geo = location_geo.get(location, _empty_geo)
        base_pops = base_location_pops.get(location, [])
        geo_vals = (geo.continent, geo.superregion, geo.region, geo.area, geo.province)
        for idx, attrs in enumerate(location_pops.get(location, [])):
            if "size" in attrs:
                attrs["size"] = normalize_size(attrs["size"])

            if idx < len(base_pops):
                base = base_pops[idx]
                differs = list(attrs.items()) != list(base.items())
                base_summary = ", ".join(f"{k}={v}" for k, v in base.items()) if differs else ""
            else:
                differs = True
                base_summary = "<missing in basegame>"

            extras = "; ".join(f"{k}={v}" for k, v in attrs.items() if k not in STD_ATTR_ORDER)

            rows.append((
                location,
                attrs.get("type", ""),
                attrs.get("size", ""),
                attrs.get("culture", ""),
                attrs.get("religion", ""),
                *geo_vals,
                extras,
                "yes" if differs else "",
                base_summary,
            ))
    return rows


# ---------------------------------------------------------------------------
# Excel writing — style objects created once and reused
# ---------------------------------------------------------------------------

_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_NORMAL_FONT  = Font(name="Arial", size=10)
_CELL_ALIGN   = Alignment(vertical="center")
_DIFFERS_FILL = PatternFill("solid", start_color="FFF1D6")
_GEO_FILL     = PatternFill("solid", start_color="EBF3FB")
_ALT_FILL     = PatternFill("solid", start_color="F8F8F8")
_BOLD_FONT    = Font(name="Arial", bold=True, size=10)
_TOTALS_FILL  = PatternFill("solid", start_color="D9E1F2")
_TOTALS_FONT  = Font(name="Arial", bold=True, size=10, color="1F4E79")


def write_excel(rows: list[tuple], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pops"

    n_rows   = len(rows)
    size_col = get_column_letter(_SIZE_COL_IDX + 1)
    last_col = get_column_letter(len(TABLE_COLUMNS))
    data_end = _DATA_START_ROW + n_rows - 1

    # --- Row 1: live totals (always visible, never filtered) ---
    label_cell = ws.cell(row=1, column=1, value="▶ FILTERED TOTAL")
    label_cell.font      = _TOTALS_FONT
    label_cell.fill      = _TOTALS_FILL
    label_cell.alignment = Alignment(horizontal="left", vertical="center")

    count_ci   = TABLE_COLUMNS.index("type") + 1
    count_cell = ws.cell(
        row=1, column=count_ci,
        value=f"=SUBTOTAL(103,{size_col}{_DATA_START_ROW}:{size_col}{data_end})",
    )
    count_cell.font          = _TOTALS_FONT
    count_cell.fill          = _TOTALS_FILL
    count_cell.number_format = '#,##0" rows"'
    count_cell.alignment     = Alignment(horizontal="right", vertical="center")

    total_cell = ws.cell(
        row=1, column=_SIZE_COL_IDX + 1,
        value=f"=SUBTOTAL(9,{size_col}{_DATA_START_ROW}:{size_col}{data_end})",
    )
    total_cell.font          = _TOTALS_FONT
    total_cell.fill          = _TOTALS_FILL
    total_cell.number_format = "#,##0.000"
    total_cell.alignment     = Alignment(horizontal="right", vertical="center")

    for ci in range(len(TABLE_COLUMNS)):
        cell = ws.cell(row=1, column=ci + 1)
        if cell.value is None:
            cell.fill = _TOTALS_FILL
    ws.row_dimensions[1].height = 18

    # --- Row 2: column headers with filter dropdowns ---
    for ci, col_name in enumerate(TABLE_COLUMNS):
        cell = ws.cell(row=2, column=ci + 1, value=col_name.title())
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(ci + 1)].width = COL_WIDTHS.get(col_name, 15)
    ws.row_dimensions[2].height = 20

    # Freeze rows 1 and 2 so both totals and headers stay visible
    ws.freeze_panes = "A3"

    # Auto-filter on the header row — only affects rows 3+ so totals row is never hidden
    ws.auto_filter.ref = f"A2:{last_col}{data_end}"

    # --- Rows 3+: data ---
    for ri, row in enumerate(rows, _DATA_START_ROW):
        is_differs = row[_DIFFERS_COL_IDX] == "yes"
        is_alt     = (ri % 2 == 0) and not is_differs
        ws.row_dimensions[ri].height = 16

        for ci, value in enumerate(row):
            if ci == _SIZE_COL_IDX and value:
                try:
                    value = float(value)
                except ValueError:
                    pass

            cell = ws.cell(row=ri, column=ci + 1, value=value)
            cell.font      = _NORMAL_FONT
            cell.alignment = _CELL_ALIGN

            if is_differs:
                cell.fill = _DIFFERS_FILL
            elif ci in _GEO_COL_IDXS:
                cell.fill = _GEO_FILL
            elif is_alt:
                cell.fill = _ALT_FILL

        ws.cell(row=ri, column=_SIZE_COL_IDX + 1).number_format = "0.000"

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    summary = [
        ("Total Pop Rows",               n_rows),
        ("Unique Locations",             len({r[0] for r in rows})),
        ("Rows Differing from Basegame", sum(1 for r in rows if r[_DIFFERS_COL_IDX] == "yes")),
        ("Total Population",             f"=SUM(Pops!{size_col}{_DATA_START_ROW}:{size_col}{data_end})"),
    ]
    for ri, (label, value) in enumerate(summary, 1):
        ws2.cell(row=ri, column=1, value=label).font = _BOLD_FONT
        ws2.cell(row=ri, column=2, value=value).font = _NORMAL_FONT
    ws2["B4"].number_format = "#,##0.000"
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16

    wb.save(output_path)
    print(f"Saved {n_rows} rows to {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    settings = load_settings()
    if not settings:
        settings = prompt_for_paths()

    base_game_path  = Path(settings["base_game"]).expanduser()
    mod_folder_path = Path(settings["mod_folder"]).expanduser()

    pops_path        = mod_folder_path / "main_menu/setup/start/06_pops.txt"
    definitions_path = base_game_path  / "game/in_game/map_data/definitions.txt"
    base_pops_path   = base_game_path  / BASE_POPS_REL

    if len(sys.argv) >= 2: pops_path        = Path(sys.argv[1])
    if len(sys.argv) >= 3: definitions_path = Path(sys.argv[2])
    if len(sys.argv) >= 4: base_pops_path   = Path(sys.argv[3])

    output_path = Path(__file__).resolve().parent / "06_pops.xlsx"
    if len(sys.argv) >= 5: output_path = Path(sys.argv[4])

    print(f"Loading definitions from {definitions_path} ...")
    location_geo = parse_definitions(definitions_path)

    print(f"Loading mod pops from {pops_path} ...")
    location_order, location_pops = parse_pops(pops_path)

    print(f"Loading basegame pops from {base_pops_path} ...")
    _, base_location_pops = parse_pops(base_pops_path)

    print("Building rows ...")
    rows = build_rows(location_order, location_pops, location_geo, base_location_pops)

    print(f"Writing Excel file to {output_path} ...")
    write_excel(rows, output_path)


if __name__ == "__main__":
    main()